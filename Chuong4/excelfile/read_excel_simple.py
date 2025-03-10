from openpyxl import load_workbook
wb = load_workbook('Products.xlsx')
print(wb.sheetnames)
ws = wb[wb.sheetnames[0]]
for row in ws.values:
    for value in row:
        print(value.center(9), end='')
    print('')

